import pandas as pd
import io
import dns.asyncresolver
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

async def verify_email_mx(email):
    """Verifies that an email's domain has a valid MX record."""
    if not email:
        return False
    try:
        domain = email.split('@')[1]
        answers = await dns.asyncresolver.resolve(domain, 'MX')
        return len(answers) > 0
    except Exception:
        return False

def generate_excel(df):
    """Generates an Excel file from a Pandas DataFrame with frozen headers and auto-adjusted columns."""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Validated Leads')
    
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active
    
    # Freeze the top row
    ws.freeze_panes = 'A2'
    
    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        # Setting a minimum of 10 and max of 60 to keep it manageable
        adjusted_length = min(max(length + 2, 10), 60)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_length
        
    final_output = io.BytesIO()
    wb.save(final_output)
    return final_output.getvalue()
